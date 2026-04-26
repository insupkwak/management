from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


UNCLAIMED_STATUSES = {
    "Draft",
    "Submitted",
    "Rejected",
    "HQ Rejected",
    "HQ Submitted",
}


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9가-힣]", "", str(value or "").strip().lower())


def find_col(headers: list[Any], candidates: list[str]) -> int | None:
    normalized_headers = [normalize_header(h) for h in headers]
    normalized_candidates = {normalize_header(c) for c in candidates}

    for idx, header in enumerate(normalized_headers):
        if header in normalized_candidates:
            return idx

    return None


def parse_year(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        return str(value.year)

    text = str(value).strip()
    if not text:
        return ""

    match = re.search(r"(20\d{2})", text)
    if match:
        return match.group(1)

    return ""


def parse_amount(value: Any) -> float:
    if value is None:
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    text = text.replace("$", "")
    text = text.replace(",", "")
    text = text.replace("USD", "")
    text = text.strip()

    try:
        return float(text)
    except Exception:
        return 0.0


def normalize_dept(value: Any) -> str:
    text = str(value or "").strip().lower()

    if text in {"crew", "crw"}:
        return "crew"

    if text in {"tech", "technical"}:
        return "tech"

    return ""


def normalize_category(value: Any) -> str:
    text = str(value or "").strip().upper()

    if text in {"OPEX", "AOR"}:
        return text

    return ""


def add_amount(bucket: dict[str, Any], key_prefix: str, amount: float) -> None:
    count_key = f"{key_prefix}_count"
    amount_key = f"{key_prefix}_amount"

    bucket[count_key] = int(bucket.get(count_key, 0) or 0) + 1
    bucket[amount_key] = float(bucket.get(amount_key, 0) or 0) + amount


def empty_cost_bucket() -> dict[str, Any]:
    return {
        "opex_actual_crew_count": 0,
        "opex_actual_crew_amount": 0,
        "opex_actual_tech_count": 0,
        "opex_actual_tech_amount": 0,
        "aor_actual_crew_count": 0,
        "aor_actual_crew_amount": 0,
        "aor_actual_tech_count": 0,
        "aor_actual_tech_amount": 0,
        "aor_unclaimed_crew_count": 0,
        "aor_unclaimed_crew_amount": 0,
        "aor_unclaimed_tech_count": 0,
        "aor_unclaimed_tech_amount": 0,
    }


def aggregate_management_cost_excel(path: str | Path) -> dict[str, dict[str, dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = list(rows[0])

    invoice_date_idx = find_col(headers, ["Invoice Date", "invoice_date", "청구일자"])
    vessel_name_idx = find_col(headers, ["Vessel Name", "vessel_name", "선박명"])
    dept_idx = find_col(headers, ["Dept", "Department", "부서"])
    category_idx = find_col(headers, ["Category", "구분"])
    cost_idx = find_col(headers, ["Cost($)", "Cost", "Amount", "금액"])
    status_idx = find_col(headers, ["Status", "상태"])

    required = {
        "Invoice Date": invoice_date_idx,
        "Vessel Name": vessel_name_idx,
        "Dept": dept_idx,
        "Category": category_idx,
        "Cost($)": cost_idx,
        "Status": status_idx,
    }

    missing = [name for name, idx in required.items() if idx is None]
    if missing:
        raise ValueError(f"관리사 비용 Excel 필수 컬럼을 찾을 수 없습니다: {', '.join(missing)}")

    result: dict[str, dict[str, dict[str, Any]]] = defaultdict(
        lambda: defaultdict(empty_cost_bucket)
    )

    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        vessel_name = str(row[vessel_name_idx] or "").strip()
        if not vessel_name:
            continue

        cost_year = parse_year(row[invoice_date_idx])
        if not cost_year:
            continue

        dept = normalize_dept(row[dept_idx])
        if dept not in {"crew", "tech"}:
            continue

        category = normalize_category(row[category_idx])
        if category not in {"OPEX", "AOR"}:
            continue

        amount = parse_amount(row[cost_idx])
        status = str(row[status_idx] or "").strip()

        bucket = result[vessel_name][cost_year]

        if category == "OPEX":
            add_amount(bucket, f"opex_actual_{dept}", amount)

        elif category == "AOR":
            add_amount(bucket, f"aor_actual_{dept}", amount)

            if status in UNCLAIMED_STATUSES:
                add_amount(bucket, f"aor_unclaimed_{dept}", amount)

    return {
        vessel_name: dict(year_map)
        for vessel_name, year_map in result.items()
    }