from __future__ import annotations

import math
import re
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from management_cost_excel import aggregate_management_cost_excel

from flask import (
    Flask,
    g,
    jsonify,
    make_response,
    render_template,
    request,
    send_from_directory,
)
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = BASE_DIR / "uploads"
REPORT_UPLOAD_DIR = UPLOAD_DIR / "reports"
DB_PATH = DATA_DIR / "vessels.db"
SCHEMA_PATH = BASE_DIR / "schema.sql"

ALLOWED_REPORT_EXTENSIONS = {
    "pdf", "jpg", "jpeg", "png", "webp", "doc", "docx", "xls", "xlsx"
}
ALLOWED_POSITION_EXTENSIONS = {"xlsx"}

VALID_REPORT_KEYS = {
    "report1_file",
    "report2_file",
    "report3_file",
    "report4_file",
    "report5_file",
    "report6_file",
    "report7_file",
    "report8_file",
}

VALID_VESSEL_TYPES = {"Tanker", "Container"}
VALID_SIRE_STATUS = {"예정", "결함조치 중", "수검완료"}
VALID_TEAM_NAMES = {"TRMT1", "TRMT2", "TRMT3 & CMT2"}
DELETE_PASSWORD = "cmt2"

COC_COUNT = 10
SIRE_COUNT = 3

app = Flask(__name__)
app.config["JSON_AS_ASCII"] = False
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024


def ensure_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(
            DB_PATH,
            timeout=30,
            check_same_thread=False
        )
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA busy_timeout = 30000")
    return g.db


@app.teardown_appcontext
def close_db(error=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def column_exists(conn: sqlite3.Connection, table_name: str, column_name: str) -> bool:
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return any(row[1] == column_name for row in rows)


def ensure_vessel_columns(conn: sqlite3.Connection) -> None:
    required_columns = [
        ("vessel_type", "TEXT NOT NULL DEFAULT 'Tanker'"),
        ("management_company", "TEXT NOT NULL DEFAULT ''"),
        ("management_supervisor", "TEXT NOT NULL DEFAULT ''"),
        ("operation_manager", "TEXT NOT NULL DEFAULT ''"),
        ("owner_supervisor", "TEXT NOT NULL DEFAULT ''"),
        ("team_name", "TEXT NOT NULL DEFAULT ''"),
        ("builder", "TEXT NOT NULL DEFAULT ''"),
        ("size", "TEXT NOT NULL DEFAULT ''"),
        ("delivery_date", "TEXT NOT NULL DEFAULT ''"),
        ("next_dry_dock", "TEXT NOT NULL DEFAULT ''"),
        ("voyage_plan", "TEXT NOT NULL DEFAULT ''"),
        ("cargo_status", "TEXT NOT NULL DEFAULT 'Ballast'"),
        ("condition_report_type", "TEXT NOT NULL DEFAULT ''"),
        ("condition_report_date", "TEXT NOT NULL DEFAULT ''"),
        ("condition_report_status", "TEXT NOT NULL DEFAULT ''"),
        ("condition_report_findings", "TEXT NOT NULL DEFAULT ''"),
        ("condition_report_open_findings", "TEXT NOT NULL DEFAULT ''"),
        ("condition_report_remark", "TEXT NOT NULL DEFAULT ''"),
    ]

    for i in range(1, 16):
        required_columns.append((f"issue_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"issue_{i}_critical", "INTEGER NOT NULL DEFAULT 0"))

    for i in range(1, 11):
        required_columns.append((f"coc_type_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"coc_summary_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"coc_due_date_{i}", "TEXT NOT NULL DEFAULT ''"))

    for i in range(1, 6):
        required_columns.append((f"sire_type_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"sire_date_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"sire_status_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"sire_findings_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"sire_open_findings_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"sire_remark_{i}", "TEXT NOT NULL DEFAULT ''"))

    for report_key in VALID_REPORT_KEYS:
        required_columns.append((report_key, "TEXT NOT NULL DEFAULT ''"))

    for col_name, col_def in required_columns:
        if not column_exists(conn, "vessels", col_name):
            conn.execute(f"ALTER TABLE vessels ADD COLUMN {col_name} {col_def}")


def ensure_management_cost_table(conn: sqlite3.Connection) -> None:
    conn.execute("""
        CREATE TABLE IF NOT EXISTS vessel_management_costs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vessel_id INTEGER NOT NULL,
            cost_year TEXT NOT NULL DEFAULT '',
            opex_contract_crew_amount TEXT NOT NULL DEFAULT '',
            opex_contract_tech_amount TEXT NOT NULL DEFAULT '',
            opex_actual_crew_count TEXT NOT NULL DEFAULT '',
            opex_actual_crew_amount TEXT NOT NULL DEFAULT '',
            opex_actual_tech_count TEXT NOT NULL DEFAULT '',
            opex_actual_tech_amount TEXT NOT NULL DEFAULT '',
            aor_actual_crew_count TEXT NOT NULL DEFAULT '',
            aor_actual_crew_amount TEXT NOT NULL DEFAULT '',
            aor_actual_tech_count TEXT NOT NULL DEFAULT '',
            aor_actual_tech_amount TEXT NOT NULL DEFAULT '',
            aor_unclaimed_crew_count TEXT NOT NULL DEFAULT '',
            aor_unclaimed_crew_amount TEXT NOT NULL DEFAULT '',
            aor_unclaimed_tech_count TEXT NOT NULL DEFAULT '',
            aor_unclaimed_tech_amount TEXT NOT NULL DEFAULT '',
            cost_remark TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(vessel_id, cost_year),
            FOREIGN KEY(vessel_id) REFERENCES vessels(id) ON DELETE CASCADE
        )
    """)


def init_db() -> None:
    ensure_dirs()
    conn = sqlite3.connect(DB_PATH, timeout=30)
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("PRAGMA busy_timeout=30000;")
        conn.execute("PRAGMA foreign_keys=ON;")
        if SCHEMA_PATH.exists():
            schema_sql = SCHEMA_PATH.read_text(encoding="utf-8")
            conn.executescript(schema_sql)
        ensure_vessel_columns(conn)
        ensure_management_cost_table(conn)
        conn.commit()
    finally:
        conn.close()


def normalize_name(name: Any) -> str:
    return str(name or "").strip().lower()


def safe_float(value: Any) -> float | None:
    try:
        if value is None or str(value).strip() == "":
            return None
        num = float(value)
        if math.isfinite(num):
            return num
    except Exception:
        return None
    return None

def parse_money(value: Any) -> float:
    text = str(value or "").strip()
    if not text:
        return 0.0

    text = text.replace("$", "").replace(",", "").strip()

    try:
        return float(text)
    except Exception:
        return 0.0


def format_money(value: Any) -> str:
    amount = parse_money(value)
    if amount == 0:
        return "$ 0"
    return f"$ {amount:,.0f}"


def format_count_amount(count_value: Any, amount_value: Any) -> str:
    count_text = str(count_value or "").strip()
    amount = format_money(amount_value)

    if count_text:
        return f"{count_text}건 / {amount}"
    return amount


def allowed_file(filename: str, allowed_extensions: set[str]) -> bool:
    if "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in allowed_extensions


def get_version() -> int:
    targets = [
        BASE_DIR / "templates" / "index.html",
        BASE_DIR / "templates" / "report.html",
        BASE_DIR / "templates" / "coc_report.html",
        BASE_DIR / "templates" / "sire_report.html",
        BASE_DIR / "templates" / "condition_report.html",
        BASE_DIR / "templates" / "management_cost_report.html",
        BASE_DIR / "static" / "css" / "style.css",
        BASE_DIR / "static" / "js" / "app.js",
        Path(__file__),
    ]
    mtimes = []
    for path in targets:
        if path.exists():
            mtimes.append(int(path.stat().st_mtime))
    return max(mtimes) if mtimes else int(datetime.now().timestamp())


def no_cache_json(payload: Any, status: int = 200):
    response = make_response(jsonify(payload), status)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


def row_to_vessel_dict(row: sqlite3.Row) -> dict[str, Any]:
    vessel = dict(row)
    if vessel.get("latitude") is None:
        vessel["latitude"] = ""
    if vessel.get("longitude") is None:
        vessel["longitude"] = ""
    return vessel


def normalize_cargo_status(value: Any) -> str:
    v = str(value or "").strip()
    if v in {"Loading", "Ballast"}:
        return v
    return "Ballast"


def normalize_vessel_type(value: Any) -> str:
    v = str(value or "").strip()
    if v in VALID_VESSEL_TYPES:
        return v
    return "Tanker"


def normalize_sire_status(value: Any) -> str:
    v = str(value or "").strip()
    if v in VALID_SIRE_STATUS:
        return v
    return ""


def parse_excel_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value

    text = str(value).strip()
    if not text:
        return None

    patterns = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    ]
    for fmt in patterns:
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            continue
    return None


def parse_degree_minute_coordinate(degree: Any, minute: Any, hemisphere: Any, coord_type: str) -> float | None:
    deg = safe_float(degree)
    minute_val = safe_float(minute)
    hemi = str(hemisphere or "").strip().upper()

    if deg is None or minute_val is None or hemi not in {"N", "S", "E", "W"}:
        return None

    value = abs(deg) + (minute_val / 60.0)

    if hemi in {"S", "W"}:
        value = -value

    if coord_type == "lat" and -90 <= value <= 90:
        return round(value, 6)
    if coord_type == "lon" and -180 <= value <= 180:
        return round(value, 6)

    return None


def dms_to_decimal(deg: float, minutes: float = 0.0, seconds: float = 0.0, direction: str | None = None) -> float:
    value = abs(deg) + (minutes / 60.0) + (seconds / 3600.0)
    if direction in {"S", "W"}:
        value = -value
    elif deg < 0:
        value = -value
    return value


def parse_single_coordinate_text(text: str, coord_type: str) -> float | None:
    if not text:
        return None

    raw = str(text).strip().upper()
    if not raw:
        return None

    raw = raw.replace("º", "°").replace("’", "'").replace("`", "'").replace("“", '"').replace("”", '"')

    try:
        num = float(raw)
        if coord_type == "lat" and -90 <= num <= 90:
            return num
        if coord_type == "lon" and -180 <= num <= 180:
            return num
    except Exception:
        pass

    direction_match = re.search(r"\b([NSEW])\b|^([NSEW])|([NSEW])$", raw)
    direction = None
    if direction_match:
        for g in direction_match.groups():
            if g:
                direction = g
                break

    nums = re.findall(r"[-+]?\d+(?:\.\d+)?", raw)
    if not nums:
        return None

    try:
        deg = float(nums[0])
        minutes = float(nums[1]) if len(nums) >= 2 else 0.0
        seconds = float(nums[2]) if len(nums) >= 3 else 0.0
        value = dms_to_decimal(deg, minutes, seconds, direction)
        if coord_type == "lat" and -90 <= value <= 90:
            return value
        if coord_type == "lon" and -180 <= value <= 180:
            return value
    except Exception:
        return None

    return None


def extract_lat_lon_from_combined_text(text: str) -> tuple[float | None, float | None]:
    if not text:
        return None, None

    raw = str(text).strip().upper()
    if not raw:
        return None, None

    raw = raw.replace("º", "°").replace("’", "'").replace("`", "'").replace("“", '"').replace("”", '"')

    directional_parts = re.findall(r"([NSEW][^NSEW/]+)", raw)
    if len(directional_parts) >= 2:
        lat = None
        lon = None
        for part in directional_parts:
            part = part.strip()
            if part.startswith(("N", "S")):
                lat = parse_single_coordinate_text(part, "lat")
            elif part.startswith(("E", "W")):
                lon = parse_single_coordinate_text(part, "lon")
        if lat is not None or lon is not None:
            return lat, lon

    if "/" in raw:
        parts = [p.strip() for p in raw.split("/") if p.strip()]
        if len(parts) >= 2:
            a, b = parts[0], parts[1]
            a_lat = parse_single_coordinate_text(a, "lat")
            a_lon = parse_single_coordinate_text(a, "lon")
            b_lat = parse_single_coordinate_text(b, "lat")
            b_lon = parse_single_coordinate_text(b, "lon")

            if a_lon is not None and b_lat is not None:
                return b_lat, a_lon
            if a_lat is not None and b_lon is not None:
                return a_lat, b_lon
            if a_lat is not None and a_lon is None and b_lon is not None:
                return a_lat, b_lon
            if a_lon is not None and b_lat is not None and b_lon is None:
                return b_lat, a_lon

    return None, None


def parse_coordinate(value: Any, coord_type: str) -> float | None:
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        num = float(value)
        if coord_type == "lat" and -90 <= num <= 90:
            return num
        if coord_type == "lon" and -180 <= num <= 180:
            return num
        return None

    text = str(value).strip()
    if not text:
        return None

    return parse_single_coordinate_text(text, coord_type)


def normalize_header(text: Any) -> str:
    return re.sub(r"[^a-z0-9가-힣]", "", str(text or "").strip().lower())


def find_header_index(headers: list[Any], candidates: list[str]) -> int | None:
    normalized = [normalize_header(h) for h in headers]
    candidate_set = {normalize_header(c) for c in candidates}
    for idx, header in enumerate(normalized):
        if header in candidate_set:
            return idx
    return None


def is_new_position_format(headers: list[Any]) -> bool:
    def cell(idx: int) -> str:
        if idx >= len(headers):
            return ""
        return str(headers[idx] or "").strip().lower()

    return (
        cell(3) == "name"
        and cell(8) == "date(lt)"
        and cell(17) == "latitude"
        and cell(18) == "latitude"
        and cell(19) == "latitude"
        and cell(20) == "longitude"
        and cell(21) == "longitude"
        and cell(22) == "longitude"
    )


def pick_latest_rows_by_vessel(ws) -> tuple[dict[str, dict[str, Any]], int, int]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, 0, 0

    def parse_with_header_row(header_row_index: int) -> tuple[dict[str, dict[str, Any]], int, int]:
        if len(rows) <= header_row_index:
            return {}, 0, 0

        headers = list(rows[header_row_index])
        data_rows = rows[header_row_index + 1:]

        if is_new_position_format(headers):
            latest_by_name: dict[str, dict[str, Any]] = {}
            total_rows = 0
            invalid_count = 0

            for row in data_rows:
                if not row:
                    continue

                total_rows += 1

                raw_name = row[3] if len(row) > 3 else None
                name = str(raw_name or "").strip()
                if not name:
                    invalid_count += 1
                    continue

                dt_value = parse_excel_datetime(row[8] if len(row) > 8 else None)

                lat = parse_degree_minute_coordinate(
                    row[17] if len(row) > 17 else None,
                    row[18] if len(row) > 18 else None,
                    row[19] if len(row) > 19 else None,
                    "lat",
                )

                lon = parse_degree_minute_coordinate(
                    row[20] if len(row) > 20 else None,
                    row[21] if len(row) > 21 else None,
                    row[22] if len(row) > 22 else None,
                    "lon",
                )

                if lat is None or lon is None:
                    invalid_count += 1
                    continue

                key = normalize_name(name)
                current = latest_by_name.get(key)

                item = {
                    "name": name,
                    "latitude": lat,
                    "longitude": lon,
                    "dt": dt_value,
                }

                if current is None:
                    latest_by_name[key] = item
                    continue

                current_dt = current.get("dt")
                if dt_value and current_dt:
                    if dt_value >= current_dt:
                        latest_by_name[key] = item
                elif dt_value and not current_dt:
                    latest_by_name[key] = item
                elif not dt_value and not current_dt:
                    latest_by_name[key] = item

            return latest_by_name, total_rows, invalid_count

        name_idx = find_header_index(headers, [
            "선명", "선박명", "ship name", "shipname", "vesselname", "vessel", "name"
        ])
        date_idx = find_header_index(headers, [
            "date", "일자", "날짜", "시간", "datetime", "updatedate", "updatetime"
        ])
        lat_idx = find_header_index(headers, ["latitude", "lat", "위도"])
        lon_idx = find_header_index(headers, ["longitude", "lon", "lng", "경도"])
        position_idx = find_header_index(headers, ["위치", "position", "pos", "location"])

        if name_idx is None:
            raise ValueError("엑셀 헤더에서 선명 또는 선박명을 찾을 수 없습니다.")

        if lat_idx is None and lon_idx is None and position_idx is None:
            raise ValueError("엑셀 헤더에서 위도/경도 또는 위치 컬럼을 찾을 수 없습니다.")

        latest_by_name: dict[str, dict[str, Any]] = {}
        total_rows = 0
        invalid_count = 0

        for row in data_rows:
            if not row:
                continue

            total_rows += 1

            raw_name = row[name_idx] if name_idx < len(row) else None
            name = str(raw_name or "").strip()
            if not name:
                invalid_count += 1
                continue

            dt_value = None
            if date_idx is not None and date_idx < len(row):
                dt_value = parse_excel_datetime(row[date_idx])

            lat = None
            lon = None

            if position_idx is not None and position_idx < len(row):
                position_raw = row[position_idx]
                if position_raw not in (None, ""):
                    lat, lon = extract_lat_lon_from_combined_text(str(position_raw))

            if lat is None and lat_idx is not None and lat_idx < len(row):
                lat = parse_coordinate(row[lat_idx], "lat")

            if lon is None and lon_idx is not None and lon_idx < len(row):
                lon = parse_coordinate(row[lon_idx], "lon")

            if (lat is None or lon is None) and lat_idx is not None and lat_idx < len(row):
                extra_lat, extra_lon = extract_lat_lon_from_combined_text(str(row[lat_idx] or ""))
                if lat is None and extra_lat is not None:
                    lat = extra_lat
                if lon is None and extra_lon is not None:
                    lon = extra_lon

            if (lat is None or lon is None) and lon_idx is not None and lon_idx < len(row):
                extra_lat, extra_lon = extract_lat_lon_from_combined_text(str(row[lon_idx] or ""))
                if lat is None and extra_lat is not None:
                    lat = extra_lat
                if lon is None and extra_lon is not None:
                    lon = extra_lon

            if lat is None or lon is None:
                invalid_count += 1
                continue

            key = normalize_name(name)
            current = latest_by_name.get(key)

            item = {
                "name": name,
                "latitude": lat,
                "longitude": lon,
                "dt": dt_value,
            }

            if current is None:
                latest_by_name[key] = item
                continue

            current_dt = current.get("dt")
            if dt_value and current_dt:
                if dt_value >= current_dt:
                    latest_by_name[key] = item
            elif dt_value and not current_dt:
                latest_by_name[key] = item
            elif not dt_value and not current_dt:
                latest_by_name[key] = item

        return latest_by_name, total_rows, invalid_count

    try:
        result = parse_with_header_row(1)
        parsed_map, total_rows, _ = result
        if parsed_map or total_rows > 0:
            return result
    except ValueError:
        pass

    return parse_with_header_row(0)


def get_all_vessels() -> list[dict[str, Any]]:
    db = get_db()
    rows = db.execute("SELECT * FROM vessels ORDER BY name COLLATE NOCASE ASC").fetchall()
    return [row_to_vessel_dict(row) for row in rows]


def get_vessel_by_name(name: str) -> dict[str, Any] | None:
    db = get_db()
    row = db.execute("""
        SELECT *
        FROM vessels
        WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))
        LIMIT 1
    """, (name,)).fetchone()
    return row_to_vessel_dict(row) if row else None


def get_management_cost_by_vessel_and_year(vessel_id: int, cost_year: str) -> dict[str, Any] | None:
    db = get_db()
    row = db.execute("""
        SELECT *
        FROM vessel_management_costs
        WHERE vessel_id = ? AND cost_year = ?
        LIMIT 1
    """, (vessel_id, cost_year)).fetchone()
    return dict(row) if row else None


def normalize_count_text(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = text.replace(",", "").strip()
    try:
        return str(int(float(text)))
    except Exception:
        return ""

def normalize_amount_text(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = text.replace("$", "").replace(",", "").strip()
    try:
        return str(int(round(float(text))))
    except Exception:
        return ""

def save_management_cost(vessel_id: int, cost_year: str, payload: dict[str, Any]) -> None:
    db = get_db()

    fields = {
        "opex_contract_crew_amount": normalize_amount_text(payload.get("opexContractCrewAmount")),
        "opex_contract_tech_amount": normalize_amount_text(payload.get("opexContractTechAmount")),
        "opex_actual_crew_count": normalize_count_text(payload.get("opexActualCrewCount")),
        "opex_actual_crew_amount": normalize_amount_text(payload.get("opexActualCrewAmount")),
        "opex_actual_tech_count": normalize_count_text(payload.get("opexActualTechCount")),
        "opex_actual_tech_amount": normalize_amount_text(payload.get("opexActualTechAmount")),
        "aor_actual_crew_count": normalize_count_text(payload.get("aorActualCrewCount")),
        "aor_actual_crew_amount": normalize_amount_text(payload.get("aorActualCrewAmount")),
        "aor_actual_tech_count": normalize_count_text(payload.get("aorActualTechCount")),
        "aor_actual_tech_amount": normalize_amount_text(payload.get("aorActualTechAmount")),
        "aor_unclaimed_crew_count": normalize_count_text(payload.get("aorUnclaimedCrewCount")),
        "aor_unclaimed_crew_amount": normalize_amount_text(payload.get("aorUnclaimedCrewAmount")),
        "aor_unclaimed_tech_count": normalize_count_text(payload.get("aorUnclaimedTechCount")),
        "aor_unclaimed_tech_amount": normalize_amount_text(payload.get("aorUnclaimedTechAmount")),
        "cost_remark": str(payload.get("costRemark", "")).strip(),
    }

    existing = db.execute("""
        SELECT id
        FROM vessel_management_costs
        WHERE vessel_id = ? AND cost_year = ?
        LIMIT 1
    """, (vessel_id, cost_year)).fetchone()

    if existing:
        set_clause = ", ".join([f"{k} = ?" for k in fields.keys()])
        values = list(fields.values()) + [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), existing["id"]]
        db.execute(
            f"UPDATE vessel_management_costs SET {set_clause}, updated_at = ? WHERE id = ?",
            values
        )
    else:
        columns = ["vessel_id", "cost_year"] + list(fields.keys()) + ["updated_at"]
        placeholders = ", ".join(["?"] * len(columns))
        values = [vessel_id, cost_year] + list(fields.values()) + [datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        db.execute(
            f"INSERT INTO vessel_management_costs ({', '.join(columns)}) VALUES ({placeholders})",
            values
        )


def remove_old_report_file_if_needed(vessel: dict[str, Any], report_key: str) -> None:
    old_name = str(vessel.get(report_key, "")).strip()
    if not old_name:
        return
    old_path = REPORT_UPLOAD_DIR / old_name
    if old_path.exists() and old_path.is_file():
        try:
            old_path.unlink()
        except Exception:
            pass


def is_coc_due_within_1_month(value: Any) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    try:
        due = datetime.strptime(text[:10], "%Y-%m-%d")
    except Exception:
        return False
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    one_month_later = today + timedelta(days=30)
    return today <= due <= one_month_later


def has_any_coc_due(vessel: dict[str, Any]) -> bool:
    for i in range(1, 11):
        if is_coc_due_within_1_month(vessel.get(f"coc_due_date_{i}", "")):
            return True
    return False


def is_due_within_1_month(value: Any) -> bool:
    return is_coc_due_within_1_month(value)


def has_sire_in_progress(vessel: dict[str, Any]) -> bool:
    for i in range(1, 4):
        if str(vessel.get(f"sire_status_{i}", "")).strip() == "결함조치 중":
            return True
    return False


def has_critical_issue(vessel: dict[str, Any]) -> bool:
    for i in range(1, 16):
        issue_text = str(vessel.get(f"issue_{i}", "")).strip()
        issue_critical = int(vessel.get(f"issue_{i}_critical") or 0)
        if issue_text and issue_critical == 1:
            return True
    return False


def apply_filter_to_vessels(vessels: list[dict[str, Any]], filter_name: str) -> list[dict[str, Any]]:
    filter_name = str(filter_name or "all").strip().lower()

    if filter_name == "vlcc":
        return [v for v in vessels if str(v.get("size", "")).strip().upper() == "VLCC"]
    if filter_name == "sireprogress":
        return [v for v in vessels if has_sire_in_progress(v)]
    if filter_name == "trmt1":
        return [v for v in vessels if str(v.get("team_name", "")).strip() == "TRMT1"]
    if filter_name == "trmt2":
        return [v for v in vessels if str(v.get("team_name", "")).strip() == "TRMT2"]
    if filter_name == "cmt2":
        return [v for v in vessels if str(v.get("team_name", "")).strip() == "TRMT3 & CMT2"]
    if filter_name == "son":
        return [v for v in vessels if str(v.get("owner_supervisor", "")).strip() == "손유석 감독"]
    if filter_name == "kim":
        return [v for v in vessels if str(v.get("owner_supervisor", "")).strip() == "김흥민 감독"]
    if filter_name == "lee":
        return [v for v in vessels if str(v.get("owner_supervisor", "")).strip() == "이창주 감독"]
    if filter_name == "coc":
        return [v for v in vessels if has_any_coc_due(v)]
    if filter_name == "critical":
        return [v for v in vessels if has_critical_issue(v)]
    return vessels


def get_management_cost_report_rows(filter_name: str, selected_year: str) -> list[dict[str, Any]]:
    db = get_db()

    base_rows = get_all_vessels()
    filtered_vessels = apply_filter_to_vessels(base_rows, filter_name)
    allowed_ids = {v["id"] for v in filtered_vessels if "id" in v}

    query = """
        SELECT
            v.id AS vessel_id,
            v.name,
            v.management_company,
            v.owner_supervisor,
            c.cost_year,
            c.opex_contract_crew_amount,
            c.opex_contract_tech_amount,
            c.opex_actual_crew_count,
            c.opex_actual_crew_amount,
            c.opex_actual_tech_count,
            c.opex_actual_tech_amount,
            c.aor_actual_crew_count,
            c.aor_actual_crew_amount,
            c.aor_actual_tech_count,
            c.aor_actual_tech_amount,
            c.aor_unclaimed_crew_count,
            c.aor_unclaimed_crew_amount,
            c.aor_unclaimed_tech_count,
            c.aor_unclaimed_tech_amount,
            c.cost_remark
        FROM vessel_management_costs c
        JOIN vessels v ON v.id = c.vessel_id
    """
    params: list[Any] = []

    if selected_year:
        query += " WHERE c.cost_year = ?"
        params.append(selected_year)

    query += " ORDER BY v.name COLLATE NOCASE ASC, c.cost_year ASC"

    rows = db.execute(query, params).fetchall()
    result = [dict(row) for row in rows]

    if allowed_ids:
        result = [row for row in result if row["vessel_id"] in allowed_ids]
    else:
        result = []

    return result



def enrich_management_cost_rows(
    rows: list[dict[str, Any]],
    selected_range: str = "전체",
    selected_view: str = "전체",
) -> list[dict[str, Any]]:
    enriched = []

    for row in rows:
        new_row = dict(row)

        new_row["opex_actual_crew_display"] = format_count_amount(
            row.get("opex_actual_crew_count"), row.get("opex_actual_crew_amount")
        )
        new_row["opex_actual_tech_display"] = format_count_amount(
            row.get("opex_actual_tech_count"), row.get("opex_actual_tech_amount")
        )
        new_row["aor_actual_crew_display"] = format_count_amount(
            row.get("aor_actual_crew_count"), row.get("aor_actual_crew_amount")
        )
        new_row["aor_actual_tech_display"] = format_count_amount(
            row.get("aor_actual_tech_count"), row.get("aor_actual_tech_amount")
        )
        new_row["aor_unclaimed_crew_display"] = format_count_amount(
            row.get("aor_unclaimed_crew_count"), row.get("aor_unclaimed_crew_amount")
        )
        new_row["aor_unclaimed_tech_display"] = format_count_amount(
            row.get("aor_unclaimed_tech_count"), row.get("aor_unclaimed_tech_amount")
        )

        total_amount = 0.0

        if selected_range in {"전체", "OPEX"}:
            if selected_view in {"전체", "Crew"}:
                total_amount += parse_money(row.get("opex_actual_crew_amount"))
            if selected_view in {"전체", "Tech"}:
                total_amount += parse_money(row.get("opex_actual_tech_amount"))

        if selected_range in {"전체", "AOR"}:
            if selected_view in {"전체", "Crew"}:
                total_amount += parse_money(row.get("aor_actual_crew_amount"))
                total_amount += parse_money(row.get("aor_unclaimed_crew_amount"))
            if selected_view in {"전체", "Tech"}:
                total_amount += parse_money(row.get("aor_actual_tech_amount"))
                total_amount += parse_money(row.get("aor_unclaimed_tech_amount"))

        new_row["total_amount_display"] = format_money(total_amount)

        enriched.append(new_row)

    return enriched



def format_usd(value):
    if value in (None, '', 0, '0'):
        return ''
    try:
        num = float(str(value).replace('$', '').replace(',', '').strip())
        if num.is_integer():
            return f"$ {int(num):,}"
        return f"$ {num:,.2f}"
    except Exception:
        return str(value)
    


@app.after_request
def add_no_cache_headers(response):
    if request.path.startswith("/api/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


@app.route("/")
def index():
    return render_template("index.html", version=get_version())


@app.route("/report")
def report_page():
    filter_name = request.args.get("filter", "all")

    vessels = get_all_vessels()
    vessels = apply_filter_to_vessels(vessels, filter_name)

    report_rows = []
    critical_vessel_names = set()

    for vessel in vessels:
        vessel_has_critical = False

        for i in range(1, 16):
            issue_text = (vessel.get(f"issue_{i}") or "").strip()
            issue_critical = int(vessel.get(f"issue_{i}_critical") or 0)

            if issue_text:
                report_rows.append({
                    "name": vessel.get("name", ""),
                    "vessel_type": vessel.get("vessel_type", ""),
                    "management_company": vessel.get("management_company", ""),
                    "owner_supervisor": vessel.get("owner_supervisor", ""),
                    "team_name": vessel.get("team_name", ""),
                    "voyage_plan": vessel.get("voyage_plan", ""),
                    "issue_no": i,
                    "issue_text": issue_text,
                    "issue_critical": issue_critical,
                })

                if issue_critical == 1:
                    vessel_has_critical = True

        if vessel_has_critical:
            critical_vessel_names.add(vessel.get("name", ""))

    return render_template(
        "report.html",
        version=get_version(),
        total_count=len(vessels),
        critical_vessel_count=len(critical_vessel_names),
        report_rows=report_rows,
        current_filter=filter_name,
    )


@app.route("/coc-report")
def coc_report():
    filter_name = request.args.get("filter", "all")

    vessels = get_all_vessels()
    vessels = apply_filter_to_vessels(vessels, filter_name)

    report_rows = []
    coc_vessel_names = set()
    due_soon_vessel_names = set()

    for vessel in vessels:
        has_coc = False

        for i in range(1, 11):
            coc_type = (vessel.get(f"coc_type_{i}") or "").strip()
            coc_summary = (vessel.get(f"coc_summary_{i}") or "").strip()
            coc_due_date = (vessel.get(f"coc_due_date_{i}") or "").strip()

            if coc_type or coc_summary or coc_due_date:
                has_coc = True
                due_soon = is_due_within_1_month(coc_due_date)
                if due_soon:
                    due_soon_vessel_names.add(vessel.get("name", ""))

                report_rows.append({
                    "name": vessel.get("name", ""),
                    "vessel_type": vessel.get("vessel_type", ""),
                    "management_company": vessel.get("management_company", ""),
                    "owner_supervisor": vessel.get("owner_supervisor", ""),
                    "team_name": vessel.get("team_name", ""),
                    "coc_no": i,
                    "coc_type": coc_type,
                    "coc_summary": coc_summary,
                    "coc_due_date": coc_due_date,
                    "is_due_soon": due_soon,
                })

        if has_coc:
            coc_vessel_names.add(vessel.get("name", ""))

    return render_template(
        "coc_report.html",
        version=get_version(),
        total_count=len(vessels),
        coc_count=len(coc_vessel_names),
        due_soon_count=len(due_soon_vessel_names),
        report_rows=report_rows,
        current_filter=filter_name,
    )


@app.route("/sire-report")
def sire_report_page():
    filter_name = request.args.get("filter", "all")

    vessels = get_all_vessels()
    vessels = apply_filter_to_vessels(vessels, filter_name)

    report_rows = []
    sire_vessel_names = set()
    sire_progress_vessel_names = set()

    for vessel in vessels:
        vessel_has_sire = False
        vessel_has_progress = False

        for i in range(1, 4):
            sire_type = str(vessel.get(f"sire_type_{i}", "")).strip()
            sire_date = str(vessel.get(f"sire_date_{i}", "")).strip()
            sire_status = str(vessel.get(f"sire_status_{i}", "")).strip()
            sire_findings = str(vessel.get(f"sire_findings_{i}", "")).strip()
            sire_open_findings = str(vessel.get(f"sire_open_findings_{i}", "")).strip()
            sire_remark = str(vessel.get(f"sire_remark_{i}", "")).strip()

            if sire_type or sire_date or sire_status or sire_findings or sire_open_findings or sire_remark:
                vessel_has_sire = True

                report_rows.append({
                    "name": vessel.get("name", ""),
                    "size": vessel.get("size") or vessel.get("vessel_type", ""),
                    "management_company": vessel.get("management_company", ""),
                    "owner_supervisor": vessel.get("owner_supervisor", ""),
                    "cargo_status": vessel.get("cargo_status", ""),
                    "sire_no": i,
                    "sire_type": sire_type,
                    "sire_date": sire_date,
                    "sire_status": sire_status,
                    "sire_findings": sire_findings,
                    "sire_open_findings": sire_open_findings,
                    "sire_remark": sire_remark,
                })

                if sire_status == "결함조치 중":
                    vessel_has_progress = True

        if vessel_has_sire:
            sire_vessel_names.add(vessel.get("name", ""))

        if vessel_has_progress:
            sire_progress_vessel_names.add(vessel.get("name", ""))

    return render_template(
        "sire_report.html",
        version=get_version(),
        total_count=len(vessels),
        sire_vessel_count=len(sire_vessel_names),
        sire_progress_count=len(sire_progress_vessel_names),
        report_rows=report_rows,
        current_filter=filter_name,
    )


@app.route("/condition-report")
def condition_report_page():
    filter_name = request.args.get("filter", "all")

    vessels = get_all_vessels()
    vessels = apply_filter_to_vessels(vessels, filter_name)

    vessels_with_condition = [
        v for v in vessels
        if str(v.get("condition_report_type", "")).strip()
        or str(v.get("condition_report_date", "")).strip()
        or str(v.get("condition_report_status", "")).strip()
        or str(v.get("condition_report_findings", "")).strip()
        or str(v.get("condition_report_open_findings", "")).strip()
        or str(v.get("condition_report_remark", "")).strip()
    ]

    condition_report_count = len(vessels_with_condition)
    condition_progress_count = sum(
        1 for v in vessels_with_condition
        if str(v.get("condition_report_status", "")).strip() == "결함조치 중"
    )

    return render_template(
        "condition_report.html",
        version=get_version(),
        vessels=vessels_with_condition,
        total_count=len(vessels),
        condition_report_count=condition_report_count,
        condition_progress_count=condition_progress_count,
        current_filter=filter_name,
    )


@app.route("/management-cost-report")
def management_cost_report_page():
    filter_name = request.args.get("filter", "all")
    selected_year = str(request.args.get("year", "")).strip()
    selected_range = str(request.args.get("range", "전체")).strip()
    selected_view = str(request.args.get("view", "전체")).strip()

    if selected_range not in {"전체", "OPEX", "AOR"}:
        selected_range = "전체"

    if selected_view not in {"전체", "Crew", "Tech"}:
        selected_view = "전체"

    base_vessels = get_all_vessels()
    filtered_vessels = apply_filter_to_vessels(base_vessels, filter_name)
    report_rows = get_management_cost_report_rows(filter_name, selected_year)
    report_rows = enrich_management_cost_rows(
        report_rows,
        selected_range=selected_range,
        selected_view=selected_view,
)

    return render_template(
        "management_cost_report.html",
        version=get_version(),
        total_count=len(filtered_vessels),
        report_count=len(report_rows),
        selected_year=selected_year,
        selected_range=selected_range,
        selected_view=selected_view,
        current_filter=filter_name,
        report_rows=report_rows,
        format_usd=format_usd,
    )


@app.route("/api/vessels", methods=["GET"])
def api_get_vessels():
    current_year = datetime.now().strftime("%Y")

    db = get_db()
    rows = db.execute(
        """
        SELECT
            v.*,
            COALESCE(NULLIF(mc.opex_contract_crew_amount, ''), 0) AS mc_opex_contract_crew_amount,
            COALESCE(NULLIF(mc.opex_contract_tech_amount, ''), 0) AS mc_opex_contract_tech_amount,
            COALESCE(NULLIF(mc.opex_actual_crew_count, ''), 0) AS mc_opex_actual_crew_count,
            COALESCE(NULLIF(mc.opex_actual_crew_amount, ''), 0) AS mc_opex_actual_crew_amount,
            COALESCE(NULLIF(mc.opex_actual_tech_count, ''), 0) AS mc_opex_actual_tech_count,
            COALESCE(NULLIF(mc.opex_actual_tech_amount, ''), 0) AS mc_opex_actual_tech_amount,
            COALESCE(NULLIF(mc.aor_actual_crew_count, ''), 0) AS mc_aor_actual_crew_count,
            COALESCE(NULLIF(mc.aor_actual_crew_amount, ''), 0) AS mc_aor_actual_crew_amount,
            COALESCE(NULLIF(mc.aor_actual_tech_count, ''), 0) AS mc_aor_actual_tech_count,
            COALESCE(NULLIF(mc.aor_actual_tech_amount, ''), 0) AS mc_aor_actual_tech_amount,
            COALESCE(NULLIF(mc.aor_unclaimed_crew_count, ''), 0) AS mc_aor_unclaimed_crew_count,
            COALESCE(NULLIF(mc.aor_unclaimed_crew_amount, ''), 0) AS mc_aor_unclaimed_crew_amount,
            COALESCE(NULLIF(mc.aor_unclaimed_tech_count, ''), 0) AS mc_aor_unclaimed_tech_count,
            COALESCE(NULLIF(mc.aor_unclaimed_tech_amount, ''), 0) AS mc_aor_unclaimed_tech_amount,
            COALESCE(mc.cost_remark, '') AS mc_cost_remark,
            COALESCE(mc.cost_year, '') AS mc_cost_year
        FROM vessels v
        LEFT JOIN vessel_management_costs mc
        ON mc.vessel_id = v.id
        AND mc.cost_year = ?
        ORDER BY v.name
        """,
        (current_year,)
    ).fetchall()


    return no_cache_json([dict(row) for row in rows])


@app.route("/api/vessel/cost", methods=["GET"])
def api_get_vessel_cost():
    vessel_name = str(request.args.get("name", "")).strip()
    cost_year = str(request.args.get("year", "")).strip()

    if not vessel_name:
        return no_cache_json({"success": False, "message": "선박명이 필요합니다."}, 400)
    if not cost_year:
        return no_cache_json({"success": False, "message": "년도가 필요합니다."}, 400)

    db = get_db()
    vessel = db.execute("""
        SELECT id
        FROM vessels
        WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))
        LIMIT 1
    """, (vessel_name,)).fetchone()

    if not vessel:
        return no_cache_json({"success": False, "message": "선박을 찾을 수 없습니다."}, 404)

    cost_row = get_management_cost_by_vessel_and_year(vessel["id"], cost_year)

    return no_cache_json({
        "success": True,
        "year": cost_year,
        "data": cost_row or {},
    })


@app.route("/api/vessel", methods=["POST"])
def api_save_single_vessel():
    payload = request.get_json(silent=True) or {}

    name = str(payload.get("name", "")).strip()
    original_name = str(payload.get("_originalName", "")).strip()

    if not name:
        return no_cache_json({"success": False, "message": "선박명이 필요합니다."}, 400)

    vessel_type = normalize_vessel_type(payload.get("vesselType"))
    management_company = str(payload.get("managementCompany", "")).strip()
    management_supervisor = str(payload.get("managementSupervisor", "")).strip()
    operation_manager = str(payload.get("operationManager", "")).strip()
    owner_supervisor = str(payload.get("ownerSupervisor", "")).strip()

    team_name = str(payload.get("teamName", "")).strip()
    if team_name and team_name not in VALID_TEAM_NAMES:
        team_name = ""

    builder = str(payload.get("builder", "")).strip()
    size = str(payload.get("size", "")).strip()
    delivery_date = str(payload.get("deliveryDate", "")).strip()
    next_dry_dock = str(payload.get("nextDryDock", "")).strip()
    voyage_plan = str(payload.get("voyagePlan", "")).strip()

    cargo_status = normalize_cargo_status(payload.get("cargoStatus"))
    if vessel_type == "Container":
        cargo_status = ""

    latitude = safe_float(payload.get("latitude"))
    longitude = safe_float(payload.get("longitude"))

    fields: dict[str, Any] = {
        "name": name,
        "vessel_type": vessel_type,
        "management_company": management_company,
        "management_supervisor": management_supervisor,
        "operation_manager": operation_manager,
        "owner_supervisor": owner_supervisor,
        "team_name": team_name,
        "builder": builder,
        "size": size,
        "delivery_date": delivery_date,
        "next_dry_dock": next_dry_dock,
        "voyage_plan": voyage_plan,
        "cargo_status": cargo_status,
    }

    for i in range(1, 16):
        fields[f"issue_{i}"] = str(payload.get(f"issue{i}", "")).strip()
        fields[f"issue_{i}_critical"] = 1 if payload.get(f"issue{i}Critical") else 0

    for i in range(1, 11):
        fields[f"coc_type_{i}"] = str(payload.get(f"cocType{i}", "")).strip()
        fields[f"coc_summary_{i}"] = str(payload.get(f"cocSummary{i}", "")).strip()
        fields[f"coc_due_date_{i}"] = str(payload.get(f"cocDueDate{i}", "")).strip()

    for i in range(1, 6):
        if i > SIRE_COUNT or vessel_type == "Container":
            fields[f"sire_type_{i}"] = ""
            fields[f"sire_date_{i}"] = ""
            fields[f"sire_status_{i}"] = ""
            fields[f"sire_findings_{i}"] = ""
            fields[f"sire_open_findings_{i}"] = ""
            fields[f"sire_remark_{i}"] = ""
        else:
            fields[f"sire_type_{i}"] = str(payload.get(f"sireType{i}", "")).strip()
            fields[f"sire_date_{i}"] = str(payload.get(f"sireDate{i}", "")).strip()
            fields[f"sire_status_{i}"] = normalize_sire_status(payload.get(f"sireStatus{i}"))
            fields[f"sire_findings_{i}"] = str(payload.get(f"sireFindings{i}", "")).strip()
            fields[f"sire_open_findings_{i}"] = str(payload.get(f"sireOpenFindings{i}", "")).strip()
            fields[f"sire_remark_{i}"] = str(payload.get(f"sireRemark{i}", "")).strip()

    fields["condition_report_type"] = str(payload.get("conditionReportType", "")).strip()
    fields["condition_report_date"] = str(payload.get("conditionReportDate", "")).strip()
    fields["condition_report_status"] = normalize_sire_status(payload.get("conditionReportStatus"))
    fields["condition_report_findings"] = str(payload.get("conditionReportFindings", "")).strip()
    fields["condition_report_open_findings"] = str(payload.get("conditionReportOpenFindings", "")).strip()
    fields["condition_report_remark"] = str(payload.get("conditionReportRemark", "")).strip()

    db = get_db()
    search_name = original_name if original_name else name
    existing = db.execute("""
        SELECT *
        FROM vessels
        WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))
        LIMIT 1
    """, (search_name,)).fetchone()

    if existing:
        existing_dict = dict(existing)
        if latitude is None:
            latitude = existing_dict["latitude"]
        if longitude is None:
            longitude = existing_dict["longitude"]

        fields["latitude"] = latitude
        fields["longitude"] = longitude
        fields["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        set_clause = ", ".join([f"{key} = ?" for key in fields.keys()])
        values = list(fields.values()) + [existing_dict["id"]]

        db.execute(f"UPDATE vessels SET {set_clause} WHERE id = ?", values)
        vessel_id = existing_dict["id"]
    else:
        fields["latitude"] = latitude
        fields["longitude"] = longitude

        columns = list(fields.keys())
        placeholders = ", ".join(["?"] * len(columns))
        column_sql = ", ".join(columns)

        cursor = db.execute(
            f"INSERT INTO vessels ({column_sql}) VALUES ({placeholders})",
            [fields[col] for col in columns],
        )
        vessel_id = cursor.lastrowid

    cost_year = str(payload.get("opexContractYear", "")).strip()

    has_cost_input = any([
        str(payload.get("opexContractCrewAmount", "")).strip(),
        str(payload.get("opexContractTechAmount", "")).strip(),
        str(payload.get("opexActualCrewCount", "")).strip(),
        str(payload.get("opexActualCrewAmount", "")).strip(),
        str(payload.get("opexActualTechCount", "")).strip(),
        str(payload.get("opexActualTechAmount", "")).strip(),
        str(payload.get("aorActualCrewCount", "")).strip(), 
        str(payload.get("aorActualCrewAmount", "")).strip(),
        str(payload.get("aorActualTechCount", "")).strip(),
        str(payload.get("aorActualTechAmount", "")).strip(),
        str(payload.get("aorUnclaimedCrewCount", "")).strip(),
        str(payload.get("aorUnclaimedCrewAmount", "")).strip(),
        str(payload.get("aorUnclaimedTechCount", "")).strip(),
        str(payload.get("aorUnclaimedTechAmount", "")).strip(),
        str(payload.get("costRemark", "")).strip(),
    ])

    try:
        db.execute(f"UPDATE vessels SET {set_clause} WHERE id = ?", values)

        if cost_year and has_cost_input:
            save_management_cost(vessel_id, cost_year, payload)

        db.commit()
        return no_cache_json({"success": True, "message": "저장되었습니다."})
    except Exception as e:
        db.rollback()
        return no_cache_json({"success": False, "message": f"저장 실패: {e}"}, 500)

@app.route("/api/vessel/delete", methods=["POST"])
def api_delete_single_vessel():
    payload = request.get_json(silent=True) or {}
    name = str(payload.get("name", "")).strip()
    password = str(payload.get("password", "")).strip()

    if not name:
        return no_cache_json({"success": False, "message": "선박명이 필요합니다."}, 400)
    if password != DELETE_PASSWORD:
        return no_cache_json({"success": False, "message": "비밀번호가 올바르지 않습니다."}, 403)

    db = get_db()
    existing = db.execute("""
        SELECT id
        FROM vessels
        WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))
        LIMIT 1
    """, (name,)).fetchone()

    if not existing:
        return no_cache_json({"success": False, "message": "선박을 찾을 수 없습니다."}, 404)

    vessel_id = existing["id"]
    db.execute("DELETE FROM vessel_management_costs WHERE vessel_id = ?", (vessel_id,))
    db.execute("DELETE FROM vessels WHERE id = ?", (vessel_id,))
    db.commit()
    return no_cache_json({"success": True, "message": "삭제되었습니다."})


@app.route("/api/upload-report", methods=["POST"])
def api_upload_report():
    vessel_name = str(request.form.get("vesselName", "")).strip()
    report_key = str(request.form.get("reportKey", "")).strip()
    file = request.files.get("file")

    if not vessel_name:
        return no_cache_json({"success": False, "message": "선박명이 필요합니다."}, 400)
    if report_key not in VALID_REPORT_KEYS:
        return no_cache_json({"success": False, "message": "유효하지 않은 Report 항목입니다."}, 400)
    if not file or not file.filename:
        return no_cache_json({"success": False, "message": "업로드할 파일이 없습니다."}, 400)
    if not allowed_file(file.filename, ALLOWED_REPORT_EXTENSIONS):
        return no_cache_json({"success": False, "message": "허용되지 않는 파일 형식입니다."}, 400)

    db = get_db()
    row = db.execute("""
        SELECT *
        FROM vessels
        WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))
        LIMIT 1
    """, (vessel_name,)).fetchone()

    if not row:
        return no_cache_json({"success": False, "message": "선박을 찾을 수 없습니다."}, 404)

    vessel = dict(row)
    ext = file.filename.rsplit(".", 1)[1].lower()
    safe_name = secure_filename(vessel_name.replace(" ", "_"))
    stored_name = f"{safe_name}_{report_key}.{ext}"

    remove_old_report_file_if_needed(vessel, report_key)

    save_path = REPORT_UPLOAD_DIR / stored_name
    file.save(save_path)

    db.execute(
        f"UPDATE vessels SET {report_key} = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (stored_name, vessel["id"]),
    )
    db.commit()

    return no_cache_json({
        "success": True,
        "message": "Report 업로드 완료",
        "filename": stored_name,
        "reportKey": report_key,
    })


@app.route("/api/upload-positions", methods=["POST"])
def api_upload_positions():
    file = request.files.get("file")

    if not file or not file.filename:
        return no_cache_json({"success": False, "message": "엑셀 파일이 없습니다."}, 400)

    if not allowed_file(file.filename, ALLOWED_POSITION_EXTENSIONS):
        return no_cache_json({"success": False, "message": "xlsx 파일만 업로드 가능합니다."}, 400)

    try:
        workbook = load_workbook(file, data_only=True)
        worksheet = workbook.active

        latest_by_name, _, _ = pick_latest_rows_by_vessel(worksheet)

        db = get_db()
        db_rows = db.execute("SELECT id, name FROM vessels").fetchall()

        db_name_map = {
            normalize_name(row["name"]): {"id": row["id"], "name": row["name"]}
            for row in db_rows
        }

        updated_count = 0
        success_name_set = set()

        for normalized_name, row in latest_by_name.items():
            existing = db_name_map.get(normalized_name)
            if not existing:
                continue

            lat = row.get("latitude")
            lon = row.get("longitude")

            if lat is None or lon is None:
                continue
            if not (-90 <= float(lat) <= 90 and -180 <= float(lon) <= 180):
                continue

            db.execute("""
                UPDATE vessels
                SET latitude = ?, longitude = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                float(lat),
                float(lon),
                existing["id"],
            ))

            updated_count += 1
            success_name_set.add(existing["name"])

        db.commit()

        not_updated_vessels = sorted([
            item["name"] for item in db_name_map.values()
            if item["name"] not in success_name_set
        ])

        return no_cache_json({
            "success": True,
            "message": "위치 업데이트 완료",
            "updatedCount": updated_count,
            "notUpdatedVessels": not_updated_vessels,
        })

    except ValueError as e:
        return no_cache_json({"success": False, "message": str(e)}, 400)
    except Exception as e:
        return no_cache_json({"success": False, "message": f"엑셀 처리 실패: {e}"}, 500)


@app.route("/api/upload-management-costs", methods=["POST"])
def upload_management_costs():
    file = request.files.get("file")
    if not file or not file.filename:
        return no_cache_json({
            "success": False,
            "message": "업로드할 파일이 없습니다."
        }, 400)

    filename = str(file.filename).lower()
    if not filename.endswith(".xlsx"):
        return no_cache_json({
            "success": False,
            "message": "xlsx 파일만 업로드 가능합니다."
        }, 400)

    temp_path = DATA_DIR / "_temp_management_cost_upload.xlsx"

    try:
        file.save(temp_path)

        aggregated = aggregate_management_cost_excel(temp_path)

        db = get_db()
        vessel_rows = db.execute("""
            SELECT id, name
            FROM vessels
        """).fetchall()

        vessel_map = {
            str(row["name"]).strip().upper(): row["id"]
            for row in vessel_rows
            if str(row["name"]).strip()
        }

        updated_count = 0
        failed_vessels: list[str] = []

        for vessel_name, year_map in aggregated.items():
            vessel_id = vessel_map.get(str(vessel_name).strip().upper())

            if not vessel_id:
                failed_vessels.append(vessel_name)
                continue

            vessel_updated = False

            for cost_year, data in year_map.items():
                existing = db.execute("""
                    SELECT id
                    FROM vessel_management_costs
                    WHERE vessel_id = ? AND cost_year = ?
                    LIMIT 1
                """, (vessel_id, cost_year)).fetchone()

                values = (
                    str(data.get("opex_actual_crew_count", 0) or 0),
                    str(int(round(float(data.get("opex_actual_crew_amount", 0) or 0)))),
                    str(data.get("opex_actual_tech_count", 0) or 0),
                    str(int(round(float(data.get("opex_actual_tech_amount", 0) or 0)))),
                    str(data.get("aor_actual_crew_count", 0) or 0),
                    str(int(round(float(data.get("aor_actual_crew_amount", 0) or 0)))),
                    str(data.get("aor_actual_tech_count", 0) or 0),
                    str(int(round(float(data.get("aor_actual_tech_amount", 0) or 0)))),
                    str(data.get("aor_unclaimed_crew_count", 0) or 0),
                    str(int(round(float(data.get("aor_unclaimed_crew_amount", 0) or 0)))),
                    str(data.get("aor_unclaimed_tech_count", 0) or 0),
                    str(int(round(float(data.get("aor_unclaimed_tech_amount", 0) or 0)))),
                )

                if existing:
                    db.execute("""
                        UPDATE vessel_management_costs
                        SET
                            opex_actual_crew_count = ?,
                            opex_actual_crew_amount = ?,
                            opex_actual_tech_count = ?,
                            opex_actual_tech_amount = ?,
                            aor_actual_crew_count = ?,
                            aor_actual_crew_amount = ?,
                            aor_actual_tech_count = ?,
                            aor_actual_tech_amount = ?,
                            aor_unclaimed_crew_count = ?,
                            aor_unclaimed_crew_amount = ?,
                            aor_unclaimed_tech_count = ?,
                            aor_unclaimed_tech_amount = ?,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE vessel_id = ? AND cost_year = ?
                    """, values + (vessel_id, cost_year))
                else:
                    db.execute("""
                        INSERT INTO vessel_management_costs (
                            vessel_id,
                            cost_year,
                            opex_contract_crew_amount,
                            opex_contract_tech_amount,
                            opex_actual_crew_count,
                            opex_actual_crew_amount,
                            opex_actual_tech_count,
                            opex_actual_tech_amount,
                            aor_actual_crew_count,
                            aor_actual_crew_amount,
                            aor_actual_tech_count,
                            aor_actual_tech_amount,
                            aor_unclaimed_crew_count,
                            aor_unclaimed_crew_amount,
                            aor_unclaimed_tech_count,
                            aor_unclaimed_tech_amount,
                            cost_remark,
                            updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    """, (
                        vessel_id,
                        str(cost_year).strip(),
                        "",   # opex_contract_crew_amount 유지용 기본값
                        "",   # opex_contract_tech_amount 유지용 기본값
                        values[0],
                        values[1],
                        values[2],
                        values[3],
                        values[4],
                        values[5],
                        values[6],
                        values[7],
                        values[8],
                        values[9],
                        values[10],
                        values[11],
                        "",   # cost_remark 기본값
                    ))

                vessel_updated = True

            if vessel_updated:
                updated_count += 1

        db.commit()

        return no_cache_json({
            "success": True,
            "message": "관리사 비용 업로드가 완료되었습니다.",
            "updated_count": updated_count,
            "failed_count": len(failed_vessels),
            "failed_vessels": failed_vessels,
        })

    except Exception as e:
        return no_cache_json({
            "success": False,
            "message": f"관리사 비용 업로드 중 오류가 발생했습니다: {e}"
        }, 500)

    finally:
        try:
            if temp_path.exists():
                temp_path.unlink()
        except Exception:
            pass


@app.route("/uploads/reports/<path:filename>")
def serve_report_file(filename: str):
    response = send_from_directory(REPORT_UPLOAD_DIR, filename, as_attachment=False)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.errorhandler(413)
def file_too_large(_error):
    return no_cache_json({"success": False, "message": "파일 용량이 너무 큽니다."}, 413)


init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8000, debug=True, use_reloader=False)