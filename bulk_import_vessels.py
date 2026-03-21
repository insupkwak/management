from __future__ import annotations

import math
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "data" / "vessels.db"
EXCEL_PATH = BASE_DIR / "data" / "vessels_bulk_upload.xlsx"

COC_COUNT = 10
SIRE_COUNT = 5
ISSUE_COUNT = 15


VALID_VESSEL_TYPES = {"Tanker", "Container"}
VALID_SIRE_STATUS = {"예정", "결함조치 중", "수검완료"}
VALID_TEAM_NAMES = {"TRMT1", "TRMT2", "TRMT3 & CMT2"}
VALID_CARGO_STATUS = {"Loading", "Ballast"}


def normalize_text_or_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text != "" else None


def normalize_date_or_none(value: Any) -> str | None:
    if value is None or str(value).strip() == "":
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    text = str(value).strip()

    for fmt in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
    ):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except Exception:
            continue

    return text


def normalize_float_or_none(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        num = float(value)
        if math.isfinite(num):
            return num
    except Exception:
        return None
    return None


def normalize_bool_int_or_none(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None

    text = str(value).strip().lower()
    if text in {"1", "y", "yes", "true", "t", "critical", "checked"}:
        return 1
    if text in {"0", "n", "no", "false", "f", "normal", "unchecked"}:
        return 0
    return 0


def normalize_vessel_type_or_none(value: Any) -> str | None:
    text = normalize_text_or_none(value)
    if text is None:
        return None
    return text if text in VALID_VESSEL_TYPES else "Tanker"


def normalize_team_name_or_none(value: Any) -> str | None:
    text = normalize_text_or_none(value)
    if text is None:
        return None
    return text if text in VALID_TEAM_NAMES else ""


def normalize_owner_supervisor_or_none(value: Any) -> str | None:
    return normalize_text_or_none(value)


def normalize_cargo_status_for_bulk(value: Any, vessel_type: str | None) -> str | None:
    if value is None or str(value).strip() == "":
        return None

    if vessel_type == "Container":
        return ""

    text = str(value).strip()
    if text in VALID_CARGO_STATUS:
        return text
    return "Ballast"


def normalize_sire_status_or_none(value: Any) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip()
    return text if text in VALID_SIRE_STATUS else ""


def build_header_map(sheet) -> dict[str, int]:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    return {header: idx for idx, header in enumerate(headers) if header}


def get_cell(row: list[Any], header_map: dict[str, int], key: str) -> Any:
    idx = header_map.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def ensure_required_columns_exist(conn: sqlite3.Connection) -> None:
    rows = conn.execute("PRAGMA table_info(vessels)").fetchall()
    existing = {row[1] for row in rows}

    required_columns: list[tuple[str, str]] = [
        ("name", "TEXT NOT NULL DEFAULT ''"),
        ("vessel_type", "TEXT NOT NULL DEFAULT 'Tanker'"),
        ("management_company", "TEXT NOT NULL DEFAULT ''"),
        ("management_supervisor", "TEXT NOT NULL DEFAULT ''"),
        ("operation_manager", "TEXT NOT NULL DEFAULT ''"),
        ("owner_supervisor", "TEXT NOT NULL DEFAULT ''"),
        ("team_name", "TEXT NOT NULL DEFAULT ''"),
        ("builder", "TEXT NOT NULL DEFAULT ''"),
        ("size", "TEXT NOT NULL DEFAULT ''"),
        ("delivery_date", "TEXT NOT NULL DEFAULT ''"),
        ("next_dry_dock", "TEXT NOT NULL DEFAULT ''"),
        ("voyage_plan", "TEXT NOT NULL DEFAULT ''"),
        ("cargo_status", "TEXT NOT NULL DEFAULT 'Ballast'"),
        ("latitude", "REAL"),
        ("longitude", "REAL"),
        ("condition_report_type", "TEXT NOT NULL DEFAULT ''"),
        ("condition_report_date", "TEXT NOT NULL DEFAULT ''"),
        ("condition_report_status", "TEXT NOT NULL DEFAULT ''"),
        ("condition_report_findings", "TEXT NOT NULL DEFAULT ''"),
        ("condition_report_open_findings", "TEXT NOT NULL DEFAULT ''"),
        ("condition_report_remark", "TEXT NOT NULL DEFAULT ''"),
    ]

    for i in range(1, ISSUE_COUNT + 1):
        required_columns.append((f"issue_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"issue_{i}_critical", "INTEGER NOT NULL DEFAULT 0"))

    for i in range(1, COC_COUNT + 1):
        required_columns.append((f"coc_type_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"coc_summary_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"coc_due_date_{i}", "TEXT NOT NULL DEFAULT ''"))

    for i in range(1, SIRE_COUNT + 1):
        required_columns.append((f"sire_type_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"sire_date_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"sire_status_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"sire_findings_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"sire_open_findings_{i}", "TEXT NOT NULL DEFAULT ''"))
        required_columns.append((f"sire_remark_{i}", "TEXT NOT NULL DEFAULT ''"))

    for col_name, col_def in required_columns:
        if col_name not in existing:
            conn.execute(f"ALTER TABLE vessels ADD COLUMN {col_name} {col_def}")

    conn.commit()


def build_record(row: list[Any], header_map: dict[str, int]) -> dict[str, Any]:
    record: dict[str, Any] = {}

    raw_name = get_cell(row, header_map, "name")
    name = str(raw_name or "").strip()
    record["name"] = name

    vessel_type = normalize_vessel_type_or_none(get_cell(row, header_map, "vessel_type"))

    def put_if_not_none(key: str, value: Any) -> None:
        if value is not None:
            record[key] = value

    put_if_not_none("vessel_type", vessel_type)
    put_if_not_none("management_company", normalize_text_or_none(get_cell(row, header_map, "management_company")))
    put_if_not_none("management_supervisor", normalize_text_or_none(get_cell(row, header_map, "management_supervisor")))
    put_if_not_none("operation_manager", normalize_text_or_none(get_cell(row, header_map, "operation_manager")))
    put_if_not_none("owner_supervisor", normalize_owner_supervisor_or_none(get_cell(row, header_map, "owner_supervisor")))
    put_if_not_none("team_name", normalize_team_name_or_none(get_cell(row, header_map, "team_name")))
    put_if_not_none("builder", normalize_text_or_none(get_cell(row, header_map, "builder")))
    put_if_not_none("size", normalize_text_or_none(get_cell(row, header_map, "size")))
    put_if_not_none("delivery_date", normalize_date_or_none(get_cell(row, header_map, "delivery_date")))
    put_if_not_none("next_dry_dock", normalize_date_or_none(get_cell(row, header_map, "next_dry_dock")))
    put_if_not_none("voyage_plan", normalize_text_or_none(get_cell(row, header_map, "voyage_plan")))
    put_if_not_none("latitude", normalize_float_or_none(get_cell(row, header_map, "latitude")))
    put_if_not_none("longitude", normalize_float_or_none(get_cell(row, header_map, "longitude")))

    cargo_status = normalize_cargo_status_for_bulk(get_cell(row, header_map, "cargo_status"), vessel_type)
    if cargo_status is not None:
        record["cargo_status"] = cargo_status

    put_if_not_none("condition_report_type", normalize_text_or_none(get_cell(row, header_map, "condition_report_type")))
    put_if_not_none("condition_report_date", normalize_date_or_none(get_cell(row, header_map, "condition_report_date")))
    put_if_not_none("condition_report_status", normalize_sire_status_or_none(get_cell(row, header_map, "condition_report_status")))
    put_if_not_none("condition_report_findings", normalize_text_or_none(get_cell(row, header_map, "condition_report_findings")))
    put_if_not_none("condition_report_open_findings", normalize_text_or_none(get_cell(row, header_map, "condition_report_open_findings")))
    put_if_not_none("condition_report_remark", normalize_text_or_none(get_cell(row, header_map, "condition_report_remark")))

    for i in range(1, ISSUE_COUNT + 1):
        put_if_not_none(f"issue_{i}", normalize_text_or_none(get_cell(row, header_map, f"issue_{i}")))
        put_if_not_none(
            f"issue_{i}_critical",
            normalize_bool_int_or_none(get_cell(row, header_map, f"issue_{i}_critical"))
        )

    for i in range(1, COC_COUNT + 1):
        put_if_not_none(f"coc_type_{i}", normalize_text_or_none(get_cell(row, header_map, f"coc_type_{i}")))
        put_if_not_none(f"coc_summary_{i}", normalize_text_or_none(get_cell(row, header_map, f"coc_summary_{i}")))
        put_if_not_none(f"coc_due_date_{i}", normalize_date_or_none(get_cell(row, header_map, f"coc_due_date_{i}")))

    for i in range(1, SIRE_COUNT + 1):
        put_if_not_none(f"sire_type_{i}", normalize_text_or_none(get_cell(row, header_map, f"sire_type_{i}")))
        put_if_not_none(f"sire_date_{i}", normalize_date_or_none(get_cell(row, header_map, f"sire_date_{i}")))
        put_if_not_none(f"sire_status_{i}", normalize_sire_status_or_none(get_cell(row, header_map, f"sire_status_{i}")))
        put_if_not_none(f"sire_findings_{i}", normalize_text_or_none(get_cell(row, header_map, f"sire_findings_{i}")))
        put_if_not_none(f"sire_open_findings_{i}", normalize_text_or_none(get_cell(row, header_map, f"sire_open_findings_{i}")))
        put_if_not_none(f"sire_remark_{i}", normalize_text_or_none(get_cell(row, header_map, f"sire_remark_{i}")))

    return record


def apply_container_rules_for_update(update_fields: dict[str, Any], existing_row: sqlite3.Row | None) -> None:
    vessel_type = update_fields.get("vessel_type")
    if vessel_type is None and existing_row is not None:
        vessel_type = existing_row["vessel_type"]

    if vessel_type == "Container":
        update_fields["cargo_status"] = ""

        update_fields["condition_report_type"] = ""
        update_fields["condition_report_date"] = ""
        update_fields["condition_report_status"] = ""
        update_fields["condition_report_findings"] = ""
        update_fields["condition_report_open_findings"] = ""
        update_fields["condition_report_remark"] = ""

        for i in range(1, SIRE_COUNT + 1):
            update_fields[f"sire_type_{i}"] = ""
            update_fields[f"sire_date_{i}"] = ""
            update_fields[f"sire_status_{i}"] = ""
            update_fields[f"sire_findings_{i}"] = ""
            update_fields[f"sire_open_findings_{i}"] = ""
            update_fields[f"sire_remark_{i}"] = ""


def upsert_vessel(conn: sqlite3.Connection, record: dict[str, Any]) -> str:
    name = str(record.get("name", "")).strip()
    if not name:
        return "skipped"

    existing = conn.execute(
        """
        SELECT *
        FROM vessels
        WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))
        LIMIT 1
        """,
        (name,),
    ).fetchone()

    if existing:
        update_fields = {k: v for k, v in record.items() if k != "name" and v is not None}

        apply_container_rules_for_update(update_fields, existing)

        if not update_fields:
            return "skipped"

        update_fields["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        set_clause = ", ".join([f"{col} = ?" for col in update_fields.keys()])
        values = [update_fields[col] for col in update_fields.keys()]
        values.append(existing["id"])

        conn.execute(
            f"""
            UPDATE vessels
            SET {set_clause}
            WHERE id = ?
            """,
            values,
        )
        return "updated"

    insert_fields = {k: v for k, v in record.items() if v is not None}

    if not insert_fields.get("name"):
        return "skipped"

    if "vessel_type" not in insert_fields:
        insert_fields["vessel_type"] = "Tanker"

    if insert_fields["vessel_type"] == "Container":
        insert_fields["cargo_status"] = ""
        insert_fields["condition_report_type"] = ""
        insert_fields["condition_report_date"] = ""
        insert_fields["condition_report_status"] = ""
        insert_fields["condition_report_findings"] = ""
        insert_fields["condition_report_open_findings"] = ""
        insert_fields["condition_report_remark"] = ""

        for i in range(1, SIRE_COUNT + 1):
            insert_fields[f"sire_type_{i}"] = ""
            insert_fields[f"sire_date_{i}"] = ""
            insert_fields[f"sire_status_{i}"] = ""
            insert_fields[f"sire_findings_{i}"] = ""
            insert_fields[f"sire_open_findings_{i}"] = ""
            insert_fields[f"sire_remark_{i}"] = ""

    columns = list(insert_fields.keys())
    placeholders = ", ".join(["?"] * len(columns))
    column_sql = ", ".join(columns)
    values = [insert_fields[col] for col in columns]

    conn.execute(
        f"""
        INSERT INTO vessels ({column_sql})
        VALUES ({placeholders})
        """,
        values,
    )
    return "inserted"


def main() -> None:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"DB 파일이 없습니다: {DB_PATH}")

    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"엑셀 파일이 없습니다: {EXCEL_PATH}")

    workbook = load_workbook(EXCEL_PATH, data_only=True)
    sheet = workbook.active
    header_map = build_header_map(sheet)

    if "name" not in header_map:
        raise ValueError("엑셀 첫 행에 name 헤더가 필요합니다.")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    try:
        ensure_required_columns_exist(conn)

        inserted = 0
        updated = 0
        skipped = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_values = list(row)
            record = build_record(row_values, header_map)

            status = upsert_vessel(conn, record)
            if status == "inserted":
                inserted += 1
            elif status == "updated":
                updated += 1
            else:
                skipped += 1

        conn.commit()

        print("대량 업로드 완료")
        print(f"- 신규 입력: {inserted}")
        print(f"- 기존 업데이트: {updated}")
        print(f"- 건너뜀: {skipped}")

    finally:
        conn.close()


if __name__ == "__main__":
    main()